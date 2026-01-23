import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

excel_path = '/home/share/results/res_tune_1.xlsx'
image_dir = '/home/share/results/figs/global/combined/adtk_hbos_vs_chatts'
output_path = '/home/share/results/res_tune_with_images.xlsx'

# Get all image files and create a mapping from point name to filename
image_files = os.listdir(image_dir)
point_to_image = {}

# Filename pattern: ..._POINT_...
# Example: 1066_global_mask_3593591_FI_11201.PV_adtk_hbos1062_chatts4.png
# Let's try to match the POINT part. 
# It seems the POINT is after the second/third underscore and before _adtk_hbos.
for f in image_files:
    if not f.endswith('.png'):
        continue
    # Regex to find the point name. 
    # Usually it's between the ID (numbers) and _adtk_hbos
    match = re.search(r'_(\d+)_([A-Za-z0-9_\-\.]+)_adtk_hbos', f)
    if match:
        point_name = match.group(2)
        point_to_image[point_name] = f
    else:
        # Fallback for other formats if any
        # 0_global_mask_1130706_PI_40201.PV_adtk_hbos0_chatts0.png
        match2 = re.search(r'mask_\d+_([A-Za-z0-9_\-\.]+)_adtk', f)
        if match2:
            point_name = match2.group(1)
            point_to_image[point_name] = f

print(f"Total points found in images: {len(point_to_image)}")

# Load excel
df = pd.read_excel(excel_path)
print(f"Total rows in excel: {len(df)}")

# Use xlsxwriter or openpyxl? 
# openpyxl is easier for inserting images into existing sheets.
wb = load_workbook(excel_path)
ws = wb.active

# Find the index of 'ChatTS微调' column
header = [cell.value for cell in ws[1]]
try:
    chatts_col_idx = header.index('ChatTS微调') + 1
    dataset_col_idx = header.index('数据集') + 1
except ValueError as e:
    print(f"Error: {e}")
    exit(1)

print(f"Target column 'ChatTS微调' at index {chatts_col_idx}")

# Set column width for the image column
ws.column_dimensions[chr(64 + chatts_col_idx)].width = 60

# Iterate through rows starting from row 2
for row_idx in range(2, ws.max_row + 1):
    point_name = ws.cell(row=row_idx, column=dataset_col_idx).value
    if point_name in point_to_image:
        img_path = os.path.join(image_dir, point_to_image[point_name])
        img = Image(img_path)
        
        # Scale image to fit (roughly)
        # Original size might be large. Let's scale it.
        # Fixed height for rows
        ws.row_dimensions[row_idx].height = 150
        
        # Scale factor (optional, adjust as needed)
        # Let's say we want height around 140
        scale = 140.0 / img.height
        img.width = int(img.width * scale)
        img.height = 140
        
        # Calculate cell name, e.g., 'C2'
        cell_name = f"{chr(64 + chatts_col_idx)}{row_idx}"
        ws.add_image(img, cell_name)
        # print(f"Inserted image for {point_name} at {cell_name}")
    else:
        print(f"Warning: No image found for point {point_name}")

wb.save(output_path)
print(f"Finished! Output saved to {output_path}")
