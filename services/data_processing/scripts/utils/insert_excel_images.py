import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Configuration
# Since the previous file was renamed to res_tune.xlsx, we use that as input
excel_path = '/home/share/results/res_images_template.xlsx' 
image_dir = '/home/share/results/figs/global/combined/adtk_hbos_vs_chatts_8b_1024_split'
# Save to a new file or overwrite
output_path = '/home/share/results/res_images_8b_1024_split.xlsx'

def insert_images():
    # 1. Map point names to image filenames
    if not os.path.exists(image_dir):
        print(f"Error: Image directory not found: {image_dir}")
        return

    image_files = os.listdir(image_dir)
    point_to_image = {}

    print(f"Scanning {len(image_files)} files in {image_dir}...")

    # Pattern matching logic
    # Tries to extract the point name (e.g., 'FI_11201.PV') from filenames like:
    # 1066_global_mask_3593591_FI_11201.PV_adtk_hbos1062_chatts4.png
    for f in image_files:
        if not f.endswith('.png'):
            continue
        
        # Regex strategy 1: Standard generated format (e.g., 1066_global_mask_3593591_FI_11201.PV_adtk_hbos1062_chatts4.png)
        match = re.search(r'_(\d+)_([A-Za-z0-9_\-\.]+)_adtk_hbos', f)
        if match:
            point_name = match.group(2)
            point_to_image[point_name] = f
        # Regex strategy 2: Fallback for different naming conventions
        elif re.search(r'mask_\d+_([A-Za-z0-9_\-\.]+)_adtk', f):
            match2 = re.search(r'mask_\d+_([A-Za-z0-9_\-\.]+)_adtk', f)
            point_name = match2.group(1)
            point_to_image[point_name] = f
        # Regex strategy 3: Simple naming format (e.g., FI_11201.PV.png)
        else:
            # Assume the filename (without .png) is the point name
            point_name = f[:-4]  # Remove .png extension
            point_to_image[point_name] = f

    print(f"Mapped {len(point_to_image)} images to points.")

    # 2. Load Excel and insert images
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return

    print(f"Loading Excel file: {excel_path}...")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Identify columns
    header = [cell.value for cell in ws[1]]
    try:
        chatts_col_idx = header.index('ChatTS微调') + 1
        dataset_col_idx = header.index('数据集') + 1
    except ValueError as e:
        print(f"Error: Could not find required columns ('ChatTS微调' or '数据集'). {e}")
        return

    print(f"Inserting images into column {chatts_col_idx} (matched against column {dataset_col_idx})...")

    # Set column width for the image column (approx 60 chars)
    ws.column_dimensions[chr(64 + chatts_col_idx)].width = 60

    inserted_count = 0
    # Iterate rows
    for row_idx in range(2, ws.max_row + 1):
        point_name = ws.cell(row=row_idx, column=dataset_col_idx).value
        
        if point_name in point_to_image:
            img_path = os.path.join(image_dir, point_to_image[point_name])
            
            try:
                img = Image(img_path)
                
                # Resize image to fit row height
                # Target height: 140px (row height set to 150)
                ws.row_dimensions[row_idx].height = 150
                scale = 140.0 / img.height
                img.width = int(img.width * scale)
                img.height = 140
                
                # Insert to cell
                cell_name = f"{chr(64 + chatts_col_idx)}{row_idx}"
                ws.add_image(img, cell_name)
                inserted_count += 1
            except Exception as e:
                print(f"Failed to insert image for {point_name}: {e}")
        else:
            # Optional: print warning for missing images
            # print(f"Warning: No image found for {point_name}")
            pass

    # 3. Save result
    print(f"Saving output to {output_path}...")
    wb.save(output_path)
    print(f"Done. Inserted {inserted_count} images.")

if __name__ == "__main__":
    insert_images()
